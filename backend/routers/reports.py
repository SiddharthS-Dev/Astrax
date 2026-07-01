"""
AstraX EB1 Control Tower – Reports Router
==========================================
Generates Excel (.xlsx) and PDF exports of accessible experiments
using pandas/openpyxl and reportlab, served via StreamingResponse.
"""

from io import BytesIO

import pandas as pd
from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.ext.asyncio import AsyncSession

from auth import get_current_active_user, require_report_access
from database import get_db
from models import RoleName, User
from routers.experiments import _get_accessible_experiments

router = APIRouter(prefix="/reports", tags=["Reports"])


def _experiments_to_dataframe(experiments) -> pd.DataFrame:
    """Convert a list of Experiment ORM objects to a flat DataFrame."""
    rows = []
    for exp in experiments:
        rows.append(
            {
                "ID": exp.id,
                "Title": exp.title,
                "Hypothesis": exp.hypothesis or "",
                "Success Criteria": exp.success_criteria or "",
                "Status": exp.status.value if exp.status else "",
                "Target End Date": (
                    exp.target_end_date.isoformat() if exp.target_end_date else ""
                ),
                "Outcome": exp.outcome or "",
                "Next Action": exp.next_action or "",
                "Owner": exp.owner.full_name if exp.owner else "",
                "Owner Email": exp.owner.email if exp.owner else "",
                "Track": exp.track.name if exp.track else "",
            }
        )
    return pd.DataFrame(rows)


# ── Excel Export ──────────────────────────────────────────────────────────
def _generate_excel(df: pd.DataFrame) -> BytesIO:
    """Render the DataFrame into an .xlsx workbook in memory."""
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Experiments")

        # Auto-fit column widths
        worksheet = writer.sheets["Experiments"]
        for col_idx, column in enumerate(df.columns, 1):
            max_len = max(
                df[column].astype(str).map(len).max(),
                len(column),
            )
            # Cap at 50 chars wide
            adjusted_width = min(max_len + 4, 50)
            worksheet.column_dimensions[
                worksheet.cell(row=1, column=col_idx).column_letter
            ].width = adjusted_width

        # Style header row
        from openpyxl.styles import Alignment, Font, PatternFill

        header_fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    buffer.seek(0)
    return buffer


# ── PDF Export ────────────────────────────────────────────────────────────
def _generate_pdf(df: pd.DataFrame) -> BytesIO:
    """Render the DataFrame into a formatted landscape PDF."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=0.4 * inch,
        rightMargin=0.4 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch,
    )

    styles = getSampleStyleSheet()
    elements = []

    # Title
    title_style = styles["Title"]
    title_style.fontSize = 16
    title_style.textColor = colors.HexColor("#1B2A4A")
    elements.append(Paragraph("AstraX EB1 – Experiment Report", title_style))
    elements.append(Spacer(1, 0.3 * inch))

    # Prepare table data – use Paragraph for wrapping
    cell_style = styles["BodyText"]
    cell_style.fontSize = 7
    cell_style.leading = 9

    header_row = [Paragraph(f"<b>{col}</b>", cell_style) for col in df.columns]
    data_rows = []
    for _, row in df.iterrows():
        data_rows.append(
            [Paragraph(str(val)[:120], cell_style) for val in row]
        )

    table_data = [header_row] + data_rows

    # Column widths – distribute across landscape A4
    page_width = landscape(A4)[0] - 0.8 * inch
    n_cols = len(df.columns)
    col_widths = [page_width / n_cols] * n_cols

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                # Header
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B2A4A")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("FONTSIZE", (0, 0), (-1, 0), 8),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
                ("TOPPADDING", (0, 0), (-1, 0), 6),
                # Body
                ("FONTSIZE", (0, 1), (-1, -1), 7),
                ("TOPPADDING", (0, 1), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 3),
                # Alternating row colours
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F4F8")]),
                # Grid
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return buffer


# ── Export Endpoint ───────────────────────────────────────────────────────
@router.get("/experiments/export")
async def export_experiments(
    format: str = Query(
        "excel",
        description="Export format: 'excel' or 'pdf'.",
        pattern="^(excel|pdf)$",
    ),
    current_user: User = Depends(require_report_access),
    db: AsyncSession = Depends(get_db),
):
    """
    Export accessible experiments as Excel or PDF.

    **Restricted to Super Admin, Executive, and Manager roles.**
    """
    experiments = await _get_accessible_experiments(current_user, db)
    df = _experiments_to_dataframe(experiments)

    if df.empty:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No experiments available for export.",
        )

    if format == "excel":
        buffer = _generate_excel(df)
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=astrax_experiments_report.xlsx"
            },
        )

    # PDF
    buffer = _generate_pdf(df)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": "attachment; filename=astrax_experiments_report.pdf"
        },
    )
